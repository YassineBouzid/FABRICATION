from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import (QCoreApplication, QObject, QRunnable, QThread,
                          QThreadPool, pyqtSignal )
from PyQt5.QtWidgets import *#QGraphicsObject,QLineEdit,QWidget,QHBoxLayout
from PyQt5 import QtCore
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QMessageBox

import socket
import os,shutil
import sys

from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as xlimg

import qrcode
from PIL import Image
import json

import time
from datetime import date
from datetime import datetime,timedelta

import win32api, win32con

from win32com import client


import threading
from threading import Thread,Event
import psutil

import sqlite3 as sq
import base64
import hashlib

####class database:
####
####    NAME_DATABASE= covert_PATH(covert_to_excel(Projet.get()))
####    NAME_TABLE = "RAPPORT_SONDAGE"
####    NAME_REPORT= "TUBE"
####    NAME_col1="PDF_SONDAGE"
####    TUBE = f"SONDAGE{}"
####    REC1 =
####    one_record=(self.TUBE,self.REC1)
####
####    def convertToBinaryData(self,filename):
####        # Convert digital data to binary format
####        with open(filename, 'rb') as file:
####            blobData = file.read()
####        return blobData
####
####    def create_atable(self,self.NAME_DATABASE,self.NAME_TABLE,self.NAME_REPORT,self.NAME_col1):
####        cnn =sq.connect(f'{self.NAME_DATABASE}.db')
####        c = cnn.cursor()
####        c.execute(f"CREATE TABLE IF NOT EXISTS {self.NAME_TABLE} ({self.NAME_REPORT} text, {self.NAME_col1} blob)")
####        cnn.commit()
####        cnn.close()
####        print("table created done!")
####
####    def insert_one_record(self.NAME_DATABASE,self.NAME_TABLE,self.one_record):
####        cnn =sq.connect(f'{self.NAME_DATABASE}.db')
####        c = cnn.cursor()
####        c.execute(f"INSERT INTO {self.NAME_TABLE} VALUES(?,?)",self.one_record)
####        cnn.commit()
####        cnn.close()
####        print("inserted done!")
####
####        create_atable(self.NAME_DATABASE,self.NAME_TABLE,self.NAME_REPORT,self.NAME_col1)
####
####        
####        
####        insert_one_record(self.NAME_DATABASE,self.NAME_TABLE,self.one_record)

class load_path:
         
    def load_element(self, element):
        self.jsn_list="INFO_FAB.json"
        try:
          openfile = open(self.jsn_list, 'r')
          jsn_conten = json.load(openfile)
        except:
            outfile = open(self.jsn_list, "w")
            if self.jsn_list=="INFO_FAB.json":
                J_dict={"projet":"Rgz-2",
                        "Nuance":"x70  MPSL 2",
                        "Diameter":"1016 mm",
                        "Epaisseur":"12,70 mm",
                        "FAB_path_pdf":r"C:\Users\YASSINE\Desktop\SCRIPTS\fabrication",
                        "PATH_PROJET":r"\\111-pc\1750",
                        "ip_address":"192.168.2.4"}
                
                
            json_object = json.dumps(J_dict, indent=2)
            outfile.write(json_object)
            jsn_conten = J_dict
        return jsn_conten[element]

P = load_path()


def convertToBinaryData(filename):
    # Convert digital data to binary format
    with open(filename, 'rb') as file:
        blobData = file.read()
    return blobData

def create_atable(NAME_DATABASE,NAME_TABLE,NAME_REPORT,NAME_col1):
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"CREATE TABLE IF NOT EXISTS {NAME_TABLE} ({NAME_REPORT} text, {NAME_col1} blob)")
    cnn.commit()
    cnn.close()
    print("table created done!")

def insert_one_record(NAME_DATABASE,NAME_TABLE,one_record):
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"INSERT INTO {NAME_TABLE} VALUES(?,?)",one_record)
    cnn.commit()
    cnn.close()
    print("inserted done!")

def select_by_pipe_name(NAME_DATABASE,NAME_TABLE, PIPENAME):# done
    
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"SELECT * FROM {NAME_TABLE} WHERE TUBE = '{PIPENAME}'")

    rows = c.fetchall()

    for row in rows:
        print(row[:3])
    return rows

def select_last_pipe_id(NAME_DATABASE,NAME_TABLE):# done
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"SELECT * FROM {NAME_TABLE}")
    last_id = len( c.fetchall()) +1
    return str(last_id)




class Ui_FABRICATION(QMainWindow):
    
    send = pyqtSignal(str)
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QtGui.QIcon('logo.png'))
        #self.setWindowIcon(QIcon('img2.png'))
        
        
        
    def setupUi(self, FABRICATION):
        FABRICATION.setObjectName("FABRICATION")
        FABRICATION.resize(1150, 570)
        self.centralwidget = QtWidgets.QWidget(FABRICATION)
        self.centralwidget.setObjectName("centralwidget")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(10, 10, 1200, 650))
        self.widget.setObjectName("widget")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.widget)
        self.gridLayout_3.setContentsMargins(10, 10, 0, 0)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.groupBox = QtWidgets.QGroupBox(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(14)
        self.groupBox.setFont(font)
        self.groupBox.setFlat(False)
        self.groupBox.setEnabled(False)
        self.groupBox.setObjectName("groupBox")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.groupBox)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.listWidget = QtWidgets.QListWidget(self.groupBox)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(18)
        self.listWidget.setFont(font)
        self.listWidget.setObjectName("listWidget")
        self.gridLayout.addWidget(self.listWidget, 0, 0, 1, 3)
        self.label_5 = QtWidgets.QLabel(self.groupBox)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.gridLayout.addWidget(self.label_5, 1, 0, 1, 1)
        
        self.lineEdit_4 = QtWidgets.QLineEdit(self.groupBox)
        #self.lineEdit_4.returnPressed.conncet(self.keyPressEvent)
        
        self.lineEdit_4.returnPressed.connect(self.recieve)
        
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(18)
        self.lineEdit_4.setFont(font)
        self.lineEdit_4.setObjectName("lineEdit_4")
        # set default message
        #self.lineEdit_4.setText("bismi Allah")
        
       
        
        self.gridLayout.addWidget(self.lineEdit_4, 1, 1, 1, 1)
        
        self.Send = QtWidgets.QPushButton(self.groupBox)#,clicked = lambda:self.keyPressEvent())

        ########
        self.Send.clicked.connect(self.recieve)
        ########
        
        font = QtGui.QFont()
        font.setPointSize(14)
        self.Send.setFont(font)
        self.Send.setObjectName("Send")
        self.gridLayout.addWidget(self.Send, 1, 2, 1, 1)
        self.verticalLayout_3.addLayout(self.gridLayout)
        self.gridLayout_3.addWidget(self.groupBox, 0, 0, 2, 1)
        self.groupBox_4 = QtWidgets.QGroupBox(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(14)
        self.groupBox_4.setFont(font)
        self.groupBox_4.setObjectName("groupBox_4")
        
        self.verticalLayout_4 = QtWidgets.QVBoxLayout(self.groupBox_4)
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.formLayout = QtWidgets.QFormLayout()
        self.formLayout.setObjectName("formLayout")
        self.label_4 = QtWidgets.QLabel(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.label_4)
        self.lineEdit_1 = QtWidgets.QLineEdit(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.lineEdit_1.setFont(font)
        self.lineEdit_1.setObjectName("lineEdit_1")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.lineEdit_1)
        
        self.label_15 = QtWidgets.QLabel(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_15.setFont(font)
        self.label_15.setObjectName("label_15")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.label_15)
        
        self.lineEdit_2 = QtWidgets.QLineEdit(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.lineEdit_2)

        self.label_9 = QtWidgets.QLabel(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.LabelRole,self.label_9)
        
        self.nemero_fab = QtWidgets.QLineEdit(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.nemero_fab.setFont(font)
        self.nemero_fab.setObjectName("nemero_fab")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.FieldRole, self.nemero_fab)
        #self.formLayout.addWidget(self.nemero_fab)

        

        self.checkBox_1 = QtWidgets.QCheckBox(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.checkBox_1.setFont(font)
        self.checkBox_1.setObjectName("checkBox_1")
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.LabelRole, self.checkBox_1)
        
###############################################################################################################################################################################################################################################################


        
        self.Send_2 = QtWidgets.QPushButton(self.groupBox_4,clicked = lambda:self.get_correct_fabrication_number())
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.Send_2.setFont(font)
        self.Send_2.setObjectName("Send_2")        
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.FieldRole, self.Send_2)

        
        self.verticalLayout_4.addLayout(self.formLayout)
        self.gridLayout_3.addWidget(self.groupBox_4, 0, 1, 1, 1)
        self.groupBox_3 = QtWidgets.QGroupBox(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(14)
        self.groupBox_3.setFont(font)
        self.groupBox_3.setObjectName("groupBox_3")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.groupBox_3)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        
        ##############################################################
        
        ##############################################################
        
        


        
        self.label_6 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.verticalLayout_2.addWidget(self.label_6)
        self.label_12 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_12.setFont(font)
        self.label_12.setObjectName("label_12")
        self.verticalLayout_2.addWidget(self.label_12)
        self.label_11 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_11.setFont(font)
        self.label_11.setObjectName("label_11")
        self.verticalLayout_2.addWidget(self.label_11)

        self.label_10 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_10.setFont(font)
        self.label_10.setObjectName("label_10")
        self.verticalLayout_2.addWidget(self.label_10)


        self.label_16 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_16.setFont(font)
        self.label_16.setObjectName("label_16")
        self.verticalLayout_2.addWidget(self.label_16)



###############################################################################################################################################################################################################################################################################################################        
        self.label_8 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.verticalLayout_2.addWidget(self.label_8)

        
###############################################################################################################################################################################################################################################################################################################        

        
        self.label_13 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_13.setFont(font)
        self.label_13.setObjectName("label_13")
        self.verticalLayout_2.addWidget(self.label_13)
        
        self.label_14 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_14.setFont(font)
        self.label_14.setObjectName("label_14")
        self.verticalLayout_2.addWidget(self.label_14)

        self.label_18 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_18.setFont(font)
        self.label_18.setObjectName("label_8")
        self.verticalLayout_2.addWidget(self.label_18)

        

        
        self.horizontalLayout.addLayout(self.verticalLayout_2)
        
        self.verticalLayout = QtWidgets.QVBoxLayout()
        
        self.verticalLayout.setObjectName("verticalLayout")


##        self.empty_lab = QtWidgets.QLabel(self.groupBox_3)
##        font = QtGui.QFont()
##        font.setFamily("Calibri")
##        font.setPointSize(16)
##        self.empty_lab.setFont(font)
##        self.empty_lab.setObjectName("empty_lab")
##        self.verticalLayout.addWidget(self.empty_lab)
        

        
        
        
        
        
        self.fournisseur = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.fournisseur.setFont(font)
        self.fournisseur.setObjectName("fournisseur")
        self.verticalLayout.addWidget(self.fournisseur)
        self.bobine = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.bobine.setFont(font)
        self.bobine.setObjectName("bobine")
        self.verticalLayout.addWidget(self.bobine)
        self.commande = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.commande.setFont(font)
        self.commande.setObjectName("commande")
        self.verticalLayout.addWidget(self.commande)
        
        self.diametre = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.diametre.setFont(font)
        self.diametre.setObjectName("diametre")
        self.verticalLayout.addWidget(self.diametre)
        self.diametre.setText(P.load_element("Diameter"))
                

        self.date = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.date.setFont(font)
        self.date.setObjectName("date")
        self.verticalLayout.addWidget(self.date)

        
        self.installation = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.installation.setFont(font)
        self.installation.setObjectName("installation")
        self.verticalLayout.addWidget(self.installation)
        
        self.coulee = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.coulee.setFont(font)
        self.coulee.setObjectName("coulee")
        self.verticalLayout.addWidget(self.coulee)
        
        self.Epaisseur = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.Epaisseur.setFont(font)
        self.Epaisseur.setObjectName("Epaisseur")
        self.verticalLayout.addWidget(self.Epaisseur)
        self.horizontalLayout.addLayout(self.verticalLayout)
        self.gridLayout_3.addWidget(self.groupBox_3, 1, 1, 1, 1)
        self.Epaisseur.setText(P.load_element("Epaisseur"))

        self.Send_3 = QtWidgets.QPushButton(self.groupBox_3,clicked = lambda:self.confirmation())
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.Send_3.setFont(font)
        self.Send_3.setObjectName("Send_3")
        self.verticalLayout.addWidget(self.Send_3)
        #self.horizontalLayout.addLayout(self.verticalLayout)
        #self.gridLayout_3.addWidget(self.groupBox_3, , 1, 1, 1)
        #☺self.Send_3.setText(P.load_element("Send_3"))


##
##     
        self.label_7 = QtWidgets.QLabel(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(12)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.gridLayout_3.addWidget(self.label_7, 2, 1, 1, 1)

        self.label_17 = QtWidgets.QLabel(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(12)
        self.label_17.setFont(font)
        self.label_17.setObjectName("label_17")
        self.gridLayout_3.addWidget(self.label_7, 2, 1, 1, 1)
        
        self.groupBox_2 = QtWidgets.QGroupBox(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(14)
        self.groupBox_2.setFont(font)
        self.groupBox_2.setFlat(False)
        #self.groupBox_2.setCheckable(True)
        self.groupBox_2.setChecked(True)
        self.groupBox_2.setObjectName("groupBox_2")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.groupBox_2)
        
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.comboBox = QtWidgets.QComboBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(12)
        self.comboBox.setFont(font)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.gridLayout_2.addWidget(self.comboBox, 0, 1, 1, 1)

        self.checkBox_2 = QtWidgets.QCheckBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.checkBox_2.setFont(font)
        self.checkBox_2.setObjectName("checkBox_2")
        self.gridLayout_2.addWidget(self.checkBox_2, 1, 1, 1, 1)
        
        #self.formLayout.setWidget(4, QtWidgets.QFormLayout.FieldRole, self.checkBox_2)
        
        self.label = QtWidgets.QLabel(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.gridLayout_2.addWidget(self.label, 0, 0, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")

        
        self.checkBox_4 = QtWidgets.QCheckBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.checkBox_4.setFont(font)
        self.checkBox_4.setObjectName("checkBox_4")
        self.horizontalLayout_2.addWidget(self.checkBox_4)

        
        
        
        self.checkBox_5 = QtWidgets.QCheckBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.checkBox_5.setFont(font)
        self.checkBox_5.setObjectName("checkBox_5")
        self.horizontalLayout_2.addWidget(self.checkBox_5)
        self.checkBox_6 = QtWidgets.QCheckBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.checkBox_6.setFont(font)
        self.checkBox_6.setObjectName("checkBox_6")
        self.horizontalLayout_2.addWidget(self.checkBox_6)
        self.gridLayout_2.addLayout(self.horizontalLayout_2, 5, 0, 1, 2)
        self.label_3 = QtWidgets.QLabel(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.gridLayout_2.addWidget(self.label_3, 3, 0, 1, 1)
        self.textEdit = QtWidgets.QTextEdit(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.textEdit.setFont(font)
        self.textEdit.setObjectName("textEdit")
        
        self.gridLayout_2.addWidget(self.textEdit, 2, 0, 1, 2)
        self.label_2 = QtWidgets.QLabel(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.gridLayout_2.addWidget(self.label_2, 1, 0, 1, 1)
        
        self.Conform = QtWidgets.QPushButton(self.groupBox_2,clicked = lambda:self.confirmation()) 

        self.Conform.setEnabled(True)
        #self.Conform.setDefault(True)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.Conform.setFont(font)
        #self.Conform.setCheckable(False)
        self.Conform.setObjectName("Conform")
        self.gridLayout_2.addWidget(self.Conform, 6, 0, 1, 2)


########################################################################################################################################################################
##        self.Conform_info = QtWidgets.QPushButton(self.groupBox_3,clicked = lambda:self.confirmation()) 
##        self.Conform_info.setEnabled(True)
##        font = QtGui.QFont()
##        font.setFamily("Calibri")
##        font.setPointSize(14)
##        self.Conform_info.setFont(font)
##        #self.Conform.setCheckable(False)
##        self.Conform_info.setObjectName("Conform_info")
##        self.verticalLayout.addWidget(self.Conform_info)








        
        self.textEdit_2 = QtWidgets.QTextEdit(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.textEdit_2.setFont(font)
        self.textEdit_2.setObjectName("textEdit_2")
        self.gridLayout_2.addWidget(self.textEdit_2, 4, 0, 1, 2)
        self.verticalLayout_5.addLayout(self.gridLayout_2)
        self.gridLayout_3.addWidget(self.groupBox_2, 0, 2, 2, 1)
        FABRICATION.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(FABRICATION)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 0, 0))
        self.menubar.setObjectName("menubar")
        FABRICATION.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(FABRICATION)
        self.statusbar.setObjectName("statusbar")
        FABRICATION.setStatusBar(self.statusbar)
        #color and styl sheet
        
        
        
        FABRICATION.setStyleSheet("background-color: rgb(0, 210, 190);")
        self.groupBox.setStyleSheet("background-color: rgb(85, 214, 243);QGroupBox { border: 1px red;}")
        self.listWidget.setStyleSheet("background-color: rgb(255, 255, 255);")
        #button colors:
        
        self.Send_2.setStyleSheet("border-radius : 5;border : 2px solid black;background-color: rgb(255,215,0)")
        self.Send.setStyleSheet("background-color: rgb(255,215,0)")
       
        
        self.comboBox.setStyleSheet("background-color: rgb(255,215,0);")
        self.Conform.setStyleSheet("border-radius : 5;border : 2px solid black;background-color: rgb(255,215,0)")
        self.Send_3.setStyleSheet("border-radius : 5;border : 2px solid black;background-color: rgb(255,215,0)")
        
        self.label_9.setStyleSheet("font: 75;")
        #self.Conform_info.setStyleSheet("background-color: rgb(255,215,0);")
                     
        self.lineEdit_4.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.groupBox_4.setStyleSheet("QGroupBox { border: 1px solid black;}")
        self.lineEdit_1.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.lineEdit_2.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.groupBox_3.setStyleSheet("QGroupBox { border: 1px solid black;}")
        self.nemero_fab.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.fournisseur.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.bobine.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.commande.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.diametre.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.installation.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.date.setStyleSheet("background-color: rgb(255, 255, 255);")
        
        self.coulee.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Epaisseur.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.groupBox_2.setStyleSheet("QGroupBox { border: 1px solid black;}")
        self.textEdit.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.textEdit_2.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.groupBox_3.setStyleSheet("QGroupBox { border: 1px solid black;}")

        self.groupBox.setEnabled(False)
        self.groupBox_2.setEnabled(False)
        self.groupBox_3.setEnabled(False)
        
        self.lineEdit_1.setFocus()
        self.checkBox_2.setChecked(True)

        
        self.lineEdit_1.returnPressed.connect(lambda:self.set_focus(0))
        self.lineEdit_2.returnPressed.connect(lambda:self.set_focus(9))
        #self.nemero_fab.returnPressed.connect(self.operator_demo)
        
        self.nemero_fab.returnPressed.connect(self.get_correct_fabrication_number)

        self.fournisseur.returnPressed.connect(lambda:self.set_focus(2))
        self.bobine.returnPressed.connect(lambda:self.set_focus(3))
        self.commande.returnPressed.connect(lambda:self.set_focus(4))
        self.diametre.returnPressed.connect(lambda:self.set_focus(5))
        self.date.returnPressed.connect(lambda:self.set_focus(6))
        self.installation.returnPressed.connect(lambda:self.set_focus(7))
        self.coulee.returnPressed.connect(lambda:self.set_focus(8))
        self.Epaisseur.returnPressed.connect(lambda:self.set_focus(9))
        
        
        
        

        self.retranslateUi(FABRICATION)
        QtCore.QMetaObject.connectSlotsByName(FABRICATION)

    def retranslateUi(self, FABRICATION):
        _translate = QtCore.QCoreApplication.translate
        FABRICATION.setWindowTitle(_translate("FABRICATION", "DEMANDE DE RECTIFICATION"))
        
        self.groupBox.setTitle(_translate("FABRICATION", "Méssagerie"))
        self.groupBox.setStyleSheet('QGroupBox:title {''subcontrol-position: top center;''padding-left: 10px;''padding-right: 10px; } QGroupBox { border: 1px red;}')
        self.label_5.setText(_translate("FABRICATION", "Message:"))
        self.Send.setText(_translate("FABRICATION", "Send"))
        self.groupBox_4.setTitle(_translate("FABRICATION", "Operateur Fabrication:"))
        self.label_4.setText(_translate("FABRICATION", "Nom et prénom:"))
        self.label_15.setText(_translate("FABRICATION", "Code Operateur:"))
        self.Send_2.setText(_translate("FABRICATION", "ENTRER"))
        self.Send_3.setText(_translate("FABRICATION", "CONFIRMER"))
        
        self.groupBox_3.setTitle(_translate("FABRICATION", "TUBE INFO: "))
        self.label_9.setText(_translate("FABRICATION", "N° FABRICATION:"))
        self.label_6.setText(_translate("FABRICATION", "FOURNISSEUR: "))
        self.label_12.setText(_translate("FABRICATION", "N° BOBINE:"))
        #•self.empty_lab.setText(_translate("FABRICATION", "    "))
        
        
        self.label_11.setText(_translate("FABRICATION", "N° DE COMMANDE:"))

        self.label_16.setText(_translate("FABRICATION", "DATE:"))
        
        self.label_10.setText(_translate("FABRICATION", "DIAMETRE:"))
        self.label_8.setText(_translate("FABRICATION", "INSTALLATION:"))
        
        
        
        self.label_13.setText(_translate("FABRICATION", "N° COULEE:"))
        self.label_14.setText(_translate("FABRICATION", "EPAISSEUR:"))
        
        self.label_18.setText(_translate("FABRICATION", "  "))
        
        self.label_7.setText(_translate("FABRICATION", "        Developped by bouzid yassine L2-RT: 2022"))
        
        self.groupBox_2.setTitle(_translate("FABRICATION", "CONFIRMATION FAB: "))

        self.comboBox.setItemText(0, _translate("FABRICATION", ""))
        self.comboBox.setItemText(1, _translate("FABRICATION", "Main d\'oeuver"))
        self.comboBox.setItemText(2, _translate("FABRICATION", "Matière"))
        self.comboBox.setItemText(3, _translate("FABRICATION", "Milieu"))
        self.comboBox.setItemText(4, _translate("FABRICATION", "Méthode"))
        self.comboBox.setItemText(5, _translate("FABRICATION", "Machine"))
        
        self.label.setText(_translate("FABRICATION", "Anlyse des causes:"))
        
        self.checkBox_1.setText(_translate("FABRICATION", "CORRIGER"))
        self.checkBox_2.setText(_translate("FABRICATION", "IMPRIMER"))
        
        self.checkBox_4.setText(_translate("FABRICATION", "Mise en oeuver"))
        self.checkBox_5.setText(_translate("FABRICATION", "Efficacité"))
        
        self.checkBox_6.setText(_translate("FABRICATION", "Cloture"))
        
        self.label_3.setText(_translate("FABRICATION", "Action corrective:"))
        self.label_2.setText(_translate("FABRICATION", "Action curative:"))
        self.Conform.setText(_translate("FABRICATION", "CONFIRMER"))
        #self.Conform_info.setText(_translate("FABRICATION", "CONFIRMER_INFO"))

    def set_focus(self, wid):
        if wid==0:self.lineEdit_2.setFocus()
        if wid==2:self.bobine.setFocus()
        if wid==3:self.commande.setFocus()
        if wid==4: self.diametre.setFocus()
        if wid==5:self.date.setFocus()
        if wid==6:self.installation.setFocus()
        if wid==7:self.coulee.setFocus()
        if wid==8:self.Epaisseur.setFocus()
        if wid==9:self.nemero_fab.setFocus()
##        if widget==10:
##                    self.lineEdit_2.setFocus()
##        if widget==11:
##                    self.lineEdit_2.setFocus()


      
##
##    def operator_demo(self):
##        excel_file_name  = f"RAPPORT_sondage_{self.get_correct_fabrication_number()}.xlsx"
##        
##        RX1_excel_report  = r"{}\RX1_SONDAGE\{}".format(P.load_element("PATH_PROJET"),excel_file_name)
##        
##        if self.lineEdit_1.text()!="" and self.lineEdit_2.text()=="":       
##            self.groupBox.setEnabled(True)
##            if os.path.isfile(RX1_excel_report):
##                self.groupBox_2.setEnabled(True)
##                self.textEdit.setFocus()
##            else:   
##                self.groupBox_3.setEnabled(True)
##                self.lineEdit_4.setFocus()
##        else:
##            self.groupBox.setEnabled(False)
##            self.groupBox_2.setEnabled(False)
##            self.groupBox_3.setEnabled(False)
##            self.nemero_fab.setFocus()
##            self.lineEdit_1.setFocus()
##            
##    def operator_demo(self):
##        
##            
##            
##            
##

            
        print("operator_demo!!")
        
  
    #recieve()
    def keyPressEvent(self, event):
        #if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Tab:
        print('enter event detected ')
        if event.key() == Qt.Key_Return:
            self.recieve()
        

    def recieve(self):
        chat.messge_recieved.connect(self.updateReceivedMessage)
        self.send.emit(str(self.lineEdit_4.text()))
        #self.listWidget.insertItem(-1,"FAB: "+self.lineEdit_4.text())
        if chat.connectetd:
            chat.s.send(str(self.lineEdit_4.text()).encode('utf-8'))
            self.listWidget.insertItem(-1,"FAB: "+str(self.lineEdit_4.text()))
            self.lineEdit_4.setText("")
        else:
            QMessageBox.about(self, "NOT CONNECTED SERVER", "RX1 SERVER IS NOT CONNECTED")
            
    def updateReceivedMessage(self,txt):
        if txt!="":
            pass
            #it = QtGui.QStandardItem("RX1: "+txt)
            #self.listWidget.addItem("RX1: "+txt)
            #self.model.appendRow(it)
            #chat.quit()

            
    def checked(self):
         if self.checkBox_4.isChecked()==True:
             self.MISE_EN_OEUVER = "Oui"
         else :
             self.MISE_EN_OEUVER = "Non"
             
         if self.checkBox_5.isChecked()==True:
             self.EFFICACITE = "Oui"
         else :
             self.EFFICACITE = "Non"
             
         if self.checkBox_6.isChecked()==True:
             self.CLOTURE = "Oui"
         else :
             self.CLOTURE = "Non"
         for proc in psutil.process_iter():
            if proc.name() == "EXCEL.EXE" or proc.name() == "AcroRd32.exe":
                proc.kill()
             
    def rq_genrator(self,string):
            logo = Image.open('logo.png')
            basewidth = 150
             
            # adjust image size
            wpercent = (basewidth/float(logo.size[0]))
            hsize = int((float(logo.size[1])*float(wpercent)))
            logo = logo.resize((basewidth, hsize), Image.ANTIALIAS)
            QRcode = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H,
                                   box_size=20,
                                   border = 0)
            QRcode.add_data(string)
            QRcode.make()
            QRcolor = 'black'
            QRimg = QRcode.make_image(fill_color=QRcolor, back_color="white").convert('RGB')
            pos = ((QRimg.size[0] - logo.size[0]) // 2,
                   (QRimg.size[1] - logo.size[1]) // 2)
            QRimg.paste(logo, pos)
            new_width  = 170
            new_height = 130
            QRimg = QRimg.resize((new_width, new_height), Image.ANTIALIAS)
            QRimg.save('img2.png')
            return QRimg


    def converting_excel_to_pdf(self,excel_path,pdf_path):
        
        app = client.Dispatch("Excel.Application")
        #app.Interactive=False
        app.Visible= False
        pdf_wb = app.Workbooks.Open(excel_path)
        pdf_wb.ActiveSheet.ExportAsFixedFormat(0,pdf_path)
        #pdf_wb.Close()
    

            
    def get_correct_fabrication_number(self):
        pipe_name =  self.nemero_fab.text()
        getname = pipe_name.upper()
        #print("getname[0]",getname[0])
        if pipe_name== "" or getname[0] not in ['A','B','C','D','E'] :
            QMessageBox.about(self, "INCORRECT FABRICATION NUMBER", "LE NEMERO DE FABRICATION EST INCORRECT!")
            self.nemero_fab.setStyleSheet("color: red; background-color: yellow")
            return 
        if len(pipe_name)==2:
            getname = (pipe_name[0]+"0"+"0"+"0"+pipe_name[1]).upper()
            
        elif len(pipe_name)==3:
            getname = (pipe_name[0]+"0"+"0"+pipe_name[1]+pipe_name[2]).upper()
                
        elif len(pipe_name)==4:
            getname = (pipe_name[0]+"0"+pipe_name[1]+pipe_name[2]+pipe_name[3]).upper()
            
        elif len(pipe_name)==5:
            getname = pipe_name.upper()
            print("getname= ",getname)
            
        elif pipe_name.find('-BIS') != -1:
            getname = pipe_name.upper()
        else:
            QMessageBox.about(self, "INCORRECT FABRICATION NUMBER", "LE NEMERO DE FABRICATION EST INCORRECT!")
            self.nemero_fab.setStyleSheet("color: red; background-color: yellow")
            return
                
        self.nemero_fab.setStyleSheet("color: black; background-color: white")
        self.nemero_fab.setText(str(getname))
        
        excel_file_name  = f"RAPPORT_sondage_{str(getname)}.xlsx"
        RX1_excel_report  = r"{}\RX1_SONDAGE\{}".format(P.load_element("PATH_PROJET"),excel_file_name)
        
        if self.lineEdit_1.text()!="" and self.lineEdit_2.text()=="":       
            self.groupBox.setEnabled(True)
            if os.path.isfile(RX1_excel_report)and self.checkBox_1.isChecked()==False:
                
                self.groupBox_2.setEnabled(True)
                self.groupBox_3.setEnabled(False)
                self.textEdit.setFocus()
                
            elif os.path.isfile(RX1_excel_report)and self.checkBox_1.isChecked()==True: 
                self.groupBox_2.setEnabled(False)
                self.groupBox_3.setEnabled(True)
                self.lineEdit_4.setFocus()
            else:
                self.groupBox_2.setEnabled(False)
                self.groupBox_3.setEnabled(True)
                self.lineEdit_4.setFocus()
        else:
            self.groupBox.setEnabled(False)
            self.groupBox_2.setEnabled(False)
            self.groupBox_3.setEnabled(False)
            self.nemero_fab.setFocus()
            self.lineEdit_1.setFocus()

            
        #♥self.operator_demo()
        #self.fournisseur.setFocus()
        
        return str(getname)


    
    def confirmation(self):      
        if self.nemero_fab.text()==""or self.nemero_fab.text().upper()[0] not in ['A','B','C','D','E'] :
            QMessageBox.about(self, "INCORRECT FABRICATION NUMBER", "FILL UP A FABRICATION NUMBER!")
            self.nemero_fab.setStyleSheet("color: red; background-color: yellow")
            return
        self.checked()
        
        REPPORT_EMPTY =True
        try:
            variable = f"""Operateur: {self.lineEdit_1.text()}
                          \nDate:       {time.strftime("%d-%m-%y--%H:%M:%S")}
                          \nID: { select_last_pipe_id(NAME_DATABASE,NAME_TABLE)}
                          \nN° Tube: {self.nemero_fab.text().upper()} """
            self.rq_genrator(f"{self.lineEdit_1.text()}\n"+ hashlib.sha256(variable.encode('utf-8')).hexdigest())

##hashlib.sha256("Operateur: "+self.lineEdit_1.text()+
##                             "\nDate: "+time.strftime("%d-%m-%y--%H:%M:%S")+
##                             "\nID: "+ select_last_pipe_id(NAME_DATABASE,NAME_TABLE)+
##                             "\nN° Tube: "+self.nemero_fab.text().upper()

            
        except Exception as e :
            print("the exception is --------->",e)
            QMessageBox.about(self, "LOGO ERROR", "LOGO NOT FOUND!")
            

        FAB_SONDAGE_PATH = r"{}\FAB_SONDAGE".format(P.load_element("FAB_path_pdf"))
        # CREATE A SONFADAGE DIRECTORY IN FABRICATION SIDE:
        if not os.path.exists(FAB_SONDAGE_PATH):
                    os.makedirs(FAB_SONDAGE_PATH)

        # CREATE A SONFADAGE DIRECTORY IN RX1 SIDE:            
        SONDAGE_PATH_RX1 = r"{}\RX1_SONDAGE".format(P.load_element("PATH_PROJET"))
        if not os.path.exists(SONDAGE_PATH_RX1):
                        os.makedirs(SONDAGE_PATH_RX1)

        pdf_file_name    = f"RAPPORT_sondage_{self.get_correct_fabrication_number()}.pdf"
        excel_file_name  = f"RAPPORT_sondage_{self.get_correct_fabrication_number()}.xlsx"
        
        RX1_excel_report  = r"{}\RX1_SONDAGE\{}".format(P.load_element("PATH_PROJET"),excel_file_name)
        print("this is the path of excel sheet: ",RX1_excel_report)
        
        FAB_excel_report = r"{}\{}".format(FAB_SONDAGE_PATH,excel_file_name)
        FAB_pdf_report   = r"{}\{}".format(FAB_SONDAGE_PATH,pdf_file_name)
                   
        
        # CHECK IF THERE IS A SONDAGE FILE IN RX1 SIDE:
        if not os.path.isfile(RX1_excel_report):
            try:
                s_wb = load_workbook('S_template.xlsx')
                s_wb.save(FAB_excel_report)
            except Exception as e:
                
                print("exception permition is denied",e)
                
        else:
            REPPORT_EMPTY =False
            try:
                s_wb = load_workbook(RX1_excel_report)
                s_wb.save(FAB_excel_report)
                
            except Exception as e:
                print("exception permition is denied",e)
                
            
        
        s_wb = load_workbook(FAB_excel_report)    
        s_ws = s_wb.active
        s_ws.print_options.horizontalCentered = True
        
        if REPPORT_EMPTY or self.checkBox_1.isChecked()==True:
            self.checkBox_2.setChecked(False)
            
            s_ws['A3'] ="Fournisseur: "+self.fournisseur.text().upper()
            s_ws['A4'] ="Date: "+ self.date.text().upper()
            s_ws['A5'] ="N° Bobine: "+self.bobine.text().upper()
            
            s_ws['E3'] ="N° de commande: "+self.commande.text().upper()
            s_ws['E4'] ="Diamètre: "+self.diametre.text().upper()
            try:
                s_ws['E5'] ="N° de fab: " + self.get_correct_fabrication_number()
            except:
                self.nemero_fab.setFocus()
                return
            
            s_ws['F3'] ="Installation: "+self.installation.text().upper()
            s_ws['F4'] ="N° coulée: "+self.coulee.text().upper()
            s_ws['F5'] ="Epaisseur : "+self.Epaisseur.text().upper()
            self.MISE_EN_OEUVER=""
            self.EFFICACITE=""
            self.CLOTURE=""

        else:
            s_ws.add_image(xlimg("img2.png"),'F39')
            
        
        
            
        s_ws['A27'] ="ANALYSE DES CAUSES:\n"+self.comboBox.currentText()
        s_ws['A30'] ="ACTION CURATIVE IMMEDIATE POUR CORRIGER LE DEFAUT:\n" +self.textEdit.toPlainText()
        s_ws['A33'] ="ACTION CORRECTIVE POUR CORRIGER LA REPARATION DU DEFAUT:\n"+self.textEdit_2.toPlainText()
        
        s_ws['A37'] = "MISE EN OEUVER:                 " + self.MISE_EN_OEUVER
        s_ws['A39'] = "EFFICACITE:                             " + self.EFFICACITE
        s_ws['A41'] = "CLOTURE:                                " + self.CLOTURE
        
        s_ws['A43'] ="DATE ET HEURE:\n"+ time.strftime("%d-%m-%y--%H:%M:%S")
        s_ws['E37'] ="NOM ET PRENOM:   "+self.lineEdit_1.text().upper()
        #s_ws['E39'] ="VISA:"+
        
            
        
        s_wb.save(FAB_excel_report)
        s_wb.close()
        os.remove("./img2.png")

        

        #c:/Users/111/Desktop/1750/Rgz-2\Rgz-2\A\A1111\A1111-D-AAC-OK
        #if not os.path.isfile(RX1_excel_report):
        try:
           shutil.copy(FAB_excel_report,RX1_excel_report)
        except:
            QMessageBox.about(self, "COPYING FILE ERROR", "EXCEL REPORT HAS NOT BEEN COPIED\n PERMISSION DENIED")
            
            

        self.converting_excel_to_pdf(FAB_excel_report,FAB_pdf_report)
        
        print("report converted!")
        # print the report
        if self.checkBox_2.isChecked()==True:
            try:
                win32api.ShellExecute(
                0,
                "print",
                FAB_pdf_report,
                None,
                ".",
                0
                )
            except:
                QMessageBox.about(self, "PRINTER ERROR", "NO PRINTER DEVICE AVAILABLE ON THIS PC")
                
        # isert into database
        try:
            create_atable(NAME_DATABASE,NAME_TABLE,NAME_REPORT,NAME_col1)
            TUBE = f"SONDAGE_{str(self.nemero_fab.text())}"
            REC1 = convertToBinaryData(FAB_pdf_report)
            one_record=(TUBE,REC1)
            insert_one_record(NAME_DATABASE,NAME_TABLE,one_record)
        except:
            QMessageBox.about(self, "DATABASE ERROR", "DATABASE IS NOT CONNECTED")
            
        
#        #shutil.move(pdf_report,FAB_SONDAGE_PATH)
        #shutil.copy(excel_report,SONDAGE_PATH_XL)
        
        
        self.textEdit.setText("")
        self.textEdit_2.setText("")
        self.fournisseur.setText("")
        self.date.setText("")
        self.bobine.setText("")
        self.commande.setText("")
        self.diametre.setText(P.load_element("Diameter"))
        self.nemero_fab.setText("")
        self.installation.setText("")
        self.coulee.setText("")
        self.Epaisseur.setText(P.load_element("Epaisseur"))
        #self.comboBox.clear()
        self.textEdit.setText("")
        self.textEdit_2.setText("")
        

        
        self.checkBox_4.setChecked(False)
        self.checkBox_5.setChecked(False)
        self.checkBox_6.setChecked(False)
        self.checkBox_2.setChecked(True)
##        for proc in psutil.process_iter():
##            if proc.name() == "EXCEL.EXE" or proc.name() == "AcroRd32.exe":
##                proc.kill()
##
##        s_wb.close()
        print("done!!")
        
        
      
            
                    
class recievethread(QThread):    
    messge_recieved = pyqtSignal(str)
    connectetd = False
    def run(self):
        while True:
            try:
                self.s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                self.s.connect((P.load_element("ip_address"), 9997))
                
                while True:
                    try:
                        print("got connected")
                        self.connectetd = True
                        msg = self.s.recv(1024)
                        self.messge_recieved.emit(str(msg.decode('utf-8')))
                        
                        ui.listWidget.insertItem(-1,"RX1: "+ str(msg.decode('utf-8')))
                        print("RX1:",msg.decode('utf-8'))
                    except Exception as ex:
                        print("waiting 1 ... ")
                        self.connectetd = False
                        break
            except Exception as ex:
                 print("waiting 2 ... ")
             
    
       

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    FABRICATION = QtWidgets.QMainWindow()
    ui = Ui_FABRICATION()
    ui.setupUi(FABRICATION)

    
    w = 1250
    h = 670
    NAME_DATABASE= "SONDAGE_"+P.load_element("projet")
    NAME_TABLE = "RAPPORT_SONDAGE"
    NAME_REPORT= "TUBE"
    NAME_col1="PDF_SONDAGE"


    
    icon = QtGui.QIcon()
    icon.addPixmap(QtGui.QPixmap("logo.ico"), QtGui.QIcon.Selected, QtGui.QIcon.On)
    FABRICATION.setWindowIcon(icon)
    FABRICATION.setGeometry(50, 50, w, h)
    #☺FABRICATION.setFixedSize(w, h)
        
    chat  = recievethread()
    chat.start()
    FABRICATION.show()
    sys.exit(app.exec_())
    
